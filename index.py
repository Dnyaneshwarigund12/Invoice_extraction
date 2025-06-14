import os
import re
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


input_folder = "input"
output_folder = "output"
os.makedirs(input_folder, exist_ok=True)
os.makedirs(output_folder, exist_ok=True)


amazon_single_line_patterns = {
    "Order Number": r"Order Number:\s*([\d-]*\s*)",
    "Order Date": r"Order Date:\s*([\d.]*\s*)",
    "Invoice Number": r"Invoice Number :?\s*([\w-]*\s*)",
    "Invoice Date": r"Invoice Date :?\s*([\d.]*\s*)",
}
amazon_multi_line_fields = {
    "Sold By": ("Sold By :?", "PAN No:"),
    "Billing Address": ("Billing Address :?", "State/UT Code:"),
    "Shipping Address": ("Shipping Address :?", "State/UT Code:"),
}


flipkart_single_line_patterns = {
    "Order ID": r"Order ID:\s*([\w\d-]*\s*)",
    "Order Date": r"Order Date:\s*([\d-]*\s*)",
    "Invoice Number": r"Invoice Number #?\s*([\w\d-]*\s*)",
    "Invoice Date": r"Invoice Date:\s*([\d-]*\s*)",
}
flipkart_multi_line_fields = {
    "Sold By": ("Sold By:?", "Ship-from Address:"),
    "Bill To": ("Bill To", "Ship To"),
    "Ship To": ("Ship To", "Keep this invoice"),
}

def get_invoice_type(pdf_path):
    """Determine if the invoice is from Amazon or Flipkart using robust patterns."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text() or ""
            
            amazon_indicators = [
                r"amazon\.in",
                r"Tax Invoice/Bill of Supply/Cash Memo",
                r"Amazon Seller Services Private Limited",
                r"Appario Retail Private Ltd",
                r"Nisreen Sales Agency",
                r"HSN:\s*\d{4}",
            ]
            
            flipkart_indicators = [
                r"Flipkart",
                r"Flipkart India Pvt Ltd",
                r"Contact Flipkart",
                r"FSN:",
                r"Thank You!",
            ]

            
            for pattern in amazon_indicators:
                if re.search(pattern, text, re.IGNORECASE):
                    return "Amazon"
            
            # Check for Flipkart
            for pattern in flipkart_indicators:
                if re.search(pattern, text, re.IGNORECASE):
                    return "Flipkart"

            # Log extracted text for debugging
            debug_path = os.path.join(output_folder, f"debug_{os.path.basename(pdf_path)}.txt")
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"Warning: Could not identify invoice type for {pdf_path}. Extracted text saved to {debug_path}")
            return "Unknown"
    except Exception as e:
        print(f"Error reading {pdf_path}: {str(e)}")
        return "Unknown"

def extract_header(text, single_line_patterns, multi_line_fields):
    """Extract header information from the invoice text."""
    header_data = {}
    for key, pattern in single_line_patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            header_data[key] = match.group(1).strip()
    for key, (start, end) in multi_line_fields.items():
        start_pos = text.find(start)
        end_pos = text.find(end, start_pos) if end else len(text)
        if start_pos != -1:
            value = text[start_pos + len(start):end_pos].strip()
            header_data[key] = value
    return header_data

def extract_tables(pdf_path):
    """Extract table data from all pages of the PDF with fallback for malformed tables."""
    headers = None
    all_rows = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    # Fallback: Try extracting text lines that resemble a table
                    text = page.extract_text()
                    if text:
                        lines = text.split("\n")
                        for i, line in enumerate(lines):
                            # Look for common table headers
                            if any(keyword in line.lower() for keyword in ["description", "qty", "hsn", "fsn", "unit price", "total"]):
                                headers = line.split()
                                headers = [str(h).replace("\n", " ").strip() if h else "Column" for h in headers]
                                # Try to extract subsequent lines as rows
                                for row_line in lines[i+1:]:
                                    if re.match(r"^\d+\s", row_line) or any(keyword in row_line.lower() for keyword in ["iphone", "headset", "bag", "award"]):
                                        row = row_line.split()
                                        if len(row) >= len(headers):
                                            row = row[:len(headers)]
                                        else:
                                            row.extend([""] * (len(headers) - len(row)))
                                        all_rows.append(row)
                                break
                for table in tables:
                    if table and len(table) > 0:
                        if not headers:
                            headers = table[0]
                            headers = [str(h).replace("\n", " ").strip() if h else "Column" for h in headers]
                        for row in table[1:]:
                            cleaned_row = [str(cell).replace("\n", " ").strip() if cell else "" for cell in row]
                            if len(cleaned_row) < len(headers):
                                cleaned_row.extend([""] * (len(headers) - len(cleaned_row)))
                            elif len(cleaned_row) > len(headers):
                                cleaned_row = cleaned_row[:len(headers)]
                            all_rows.append(cleaned_row)
    except Exception as e:
        print(f"Error extracting tables from {pdf_path}: {str(e)}")
    return headers, all_rows

def main():
    """Main function to process invoices and write to Excel."""
    wb = Workbook()
    default_ws = wb.active
    default_ws.title = "Summary"
    default_ws.cell(row=1, column=1, value="Invoice Processing Summary")
    row = 2

    pdf_files = [
        "Iphoneinvoicev2.pdf", "1.1.pdf", "award_1.pdf", "Bag invoice main (1).pdf",  # Amazon
        "OD333957328941392100-4.pdf", "OD330090353912332100.pdf", "OD334595557473718100.pdf", "OD332423168587976100.pdf"  # Flipkart
    ]
    pdf_paths = [os.path.join(input_folder, pdf) for pdf in pdf_files]

    for pdf_path in pdf_paths:
        if not os.path.exists(pdf_path):
            print(f"Error: File {pdf_path} does not exist")
            default_ws.cell(row=row, column=1, value=f"Error: File {os.path.basename(pdf_path)} does not exist")
            row += 1
            continue

        invoice_type = get_invoice_type(pdf_path)
        if invoice_type == "Unknown":
            default_ws.cell(row=row, column=1, value=f"Error: Could not identify invoice type for {os.path.basename(pdf_path)}")
            row += 1
            continue

        try:
            with pdfplumber.open(pdf_path) as pdf:
                first_page_text = pdf.pages[0].extract_text() or ""
                if invoice_type == "Amazon":
                    header_data = extract_header(first_page_text, amazon_single_line_patterns, amazon_multi_line_fields)
                elif invoice_type == "Flipkart":
                    header_data = extract_header(first_page_text, flipkart_single_line_patterns, flipkart_multi_line_fields)

                headers, all_rows = extract_tables(pdf_path)
                sheet_name = os.path.splitext(os.path.basename(pdf_path))[0][:31]
                ws = wb.create_sheet(sheet_name)

                
                header_row = 1
                for key, value in header_data.items():
                    ws.cell(row=header_row, column=1, value=key)
                    ws.cell(row=header_row, column=2, value=value)
                    header_row += 1

                
                ws.append([])

                
                if headers and all_rows:
                    df = pd.DataFrame(all_rows, columns=headers)
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    default_ws.cell(row=row, column=1, value=f"Processed: {os.path.basename(pdf_path)} ({invoice_type})")
                else:
                    ws.cell(row=header_row + 1, column=1, value="No table data available")
                    default_ws.cell(row=row, column=1, value=f"No table data found in {os.path.basename(pdf_path)}")
                row += 1
        except Exception as e:
            print(f"Error processing {pdf_path}: {str(e)}")
            default_ws.cell(row=row, column=1, value=f"Error processing {os.path.basename(pdf_path)}: {str(e)}")
            row += 1

    
    output_path = os.path.join(output_folder, "invoices.xlsx")
    wb.save(output_path)
    print(f"Data extracted and saved to {output_path}")

if __name__ == "__main__":
    main()